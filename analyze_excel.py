import openpyxl
from openpyxl.utils import get_column_letter
import json
import pandas as pd
import sys
import io

# Set UTF-8 encoding for console output
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Load the workbook
file_path = r"C:\Users\jahub\Downloads\Project CalculEat Sheet.xlsx"
workbook = openpyxl.load_workbook(file_path, data_only=False)

print("=" * 80)
print("EXCEL FILE ANALYSIS: Project CalculEat Sheet.xlsx")
print("=" * 80)

# Get all sheet names
sheet_names = workbook.sheetnames
print(f"\nTotal sheets found: {len(sheet_names)}")
print(f"Sheet names: {sheet_names}\n")

# Expected sheets from CSV files
expected_sheets = [
    "Profile",
    "Today",
    "Yesterday",
    "New Item",
    "Items",
    "Recipe Calculator",
    "Recipes",
    "Meals",
    "History",
    "Day Base",
    "Formulas & Info"
]

print("=" * 80)
print("SHEET VERIFICATION")
print("=" * 80)
for sheet in expected_sheets:
    status = "✓ FOUND" if sheet in sheet_names else "✗ MISSING"
    print(f"{status}: {sheet}")

# Additional sheets not in CSV list
extra_sheets = [s for s in sheet_names if s not in expected_sheets]
if extra_sheets:
    print(f"\n⚠ Extra sheets found: {extra_sheets}")

print("\n" + "=" * 80)
print("DETAILED SHEET ANALYSIS")
print("=" * 80)

# Analyze each sheet
analysis_results = {}

for sheet_name in sheet_names:
    print(f"\n{'=' * 80}")
    print(f"SHEET: {sheet_name}")
    print(f"{'=' * 80}")

    sheet = workbook[sheet_name]

    # Get dimensions
    max_row = sheet.max_row
    max_col = sheet.max_column

    print(f"Dimensions: {max_row} rows × {max_col} columns")
    print(f"Range: A1 to {get_column_letter(max_col)}{max_row}")

    # Initialize sheet analysis
    sheet_data = {
        "dimensions": {"rows": max_row, "cols": max_col},
        "formulas": [],
        "merged_cells": [],
        "named_ranges": [],
        "data_validation": [],
        "conditional_formatting": [],
        "cell_data": {}
    }

    # Get merged cells
    merged = list(sheet.merged_cells.ranges)
    if merged:
        print(f"\nMerged cells: {len(merged)}")
        for merge_range in merged:
            sheet_data["merged_cells"].append(str(merge_range))
            print(f"  {merge_range}")

    # Find all formulas
    print(f"\nFormulas found:")
    formula_count = 0
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col), 1):
        for col_idx, cell in enumerate(row, 1):
            cell_ref = f"{get_column_letter(col_idx)}{row_idx}"

            # Store cell data
            cell_info = {
                "value": cell.value,
                "data_type": str(cell.data_type),
                "number_format": cell.number_format,
            }

            # Check for formula
            if cell.data_type == 'f':
                formula_count += 1
                formula_info = {
                    "cell": cell_ref,
                    "formula": cell.value,
                    "calculated_value": None
                }
                sheet_data["formulas"].append(formula_info)
                print(f"  {cell_ref}: {cell.value}")
                cell_info["formula"] = cell.value

            sheet_data["cell_data"][cell_ref] = cell_info

    # Check for data validation rules at sheet level
    if hasattr(sheet, 'data_validations') and sheet.data_validations:
        for dv in sheet.data_validations.dataValidation:
            if dv.sqref:
                validation_info = {
                    "cells": str(dv.sqref),
                    "type": dv.type,
                    "formula1": dv.formula1,
                    "formula2": dv.formula2
                }
                sheet_data["data_validation"].append(validation_info)

    if formula_count == 0:
        print("  No formulas found")
    else:
        print(f"  Total: {formula_count} formulas")

    # Check for data validation rules
    data_val_count = len(sheet_data["data_validation"])
    if data_val_count > 0:
        print(f"\nData validation rules: {data_val_count}")
        for val in sheet_data["data_validation"][:5]:  # Show first 5
            print(f"  {val['cells']}: {val['type']} - {val['formula1']}")
        if data_val_count > 5:
            print(f"  ... and {data_val_count - 5} more")

    # Print sample data (first 10 rows, first 10 cols)
    print(f"\nSample data (first 10 rows × 10 columns):")
    print("-" * 80)

    sample_rows = min(10, max_row)
    sample_cols = min(10, max_col)

    for row_idx in range(1, sample_rows + 1):
        row_data = []
        for col_idx in range(1, sample_cols + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            value = cell.value

            # Format value for display
            if value is None:
                display_value = ""
            elif cell.data_type == 'f':
                display_value = f"={value}"
            elif isinstance(value, (int, float)):
                display_value = str(value)
            else:
                display_value = str(value)[:15]  # Truncate long strings

            row_data.append(display_value)

        # Print row with cell references
        row_str = " | ".join(f"{val:>15}" for val in row_data)
        print(f"Row {row_idx:2d}: {row_str}")

    analysis_results[sheet_name] = sheet_data

# Check for named ranges (workbook level)
print("\n" + "=" * 80)
print("NAMED RANGES (Workbook Level)")
print("=" * 80)
if workbook.defined_names:
    for name, defn in workbook.defined_names.items():
        print(f"  {name}: {defn.attr_text}")
else:
    print("  No named ranges found")

# Save detailed analysis to JSON
output_file = r"c:\Users\jahub\Documents\CalculEat\excel_analysis_detailed.json"
with open(output_file, 'w', encoding='utf-8') as f:
    json.dump(analysis_results, f, indent=2, ensure_ascii=False, default=str)

print("\n" + "=" * 80)
print(f"✓ Detailed analysis saved to: {output_file}")
print("=" * 80)

# Create summary report
print("\n" + "=" * 80)
print("SUMMARY REPORT")
print("=" * 80)

total_formulas = sum(len(data["formulas"]) for data in analysis_results.values())
total_validations = sum(len(data["data_validation"]) for data in analysis_results.values())
total_merged = sum(len(data["merged_cells"]) for data in analysis_results.values())

print(f"\nTotal sheets: {len(sheet_names)}")
print(f"Total formulas: {total_formulas}")
print(f"Total data validations: {total_validations}")
print(f"Total merged cells: {total_merged}")

print("\nFormulas per sheet:")
for sheet_name, data in analysis_results.items():
    formula_count = len(data["formulas"])
    if formula_count > 0:
        print(f"  {sheet_name}: {formula_count} formulas")

print("\n" + "=" * 80)
print("ANALYSIS COMPLETE")
print("=" * 80)
